"""
✈ ALERTA DE PASSAGENS — SALVADOR (SSA)
Roda diariamente, gera planilha .xlsx e envia por e-mail.
"""

import os
import json
import smtplib
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
GEMINI_API_KEY = os.environ["GEMINI_API_KEY"]
GMAIL_USER     = os.environ["GMAIL_USER"]       # vinicyusarro@gmail.com
GMAIL_PASSWORD = os.environ["GMAIL_APP_PASSWORD"]  # App Password do Gmail
TO_EMAILS      = ["vinicyusarro@gmail.com", "lais_domitilo@hotmail.com"]
MODEL          = "gemini-2.0-flash"

# ── PROMPT ────────────────────────────────────────────────────────────────────
def build_prompt():
    today = datetime.date.today()
    d30   = today + datetime.timedelta(days=30)
    d60   = today + datetime.timedelta(days=60)
    d90   = today + datetime.timedelta(days=90)

    return f"""Você é um planejador de viagens especializado em passagens aéreas baratas no Brasil.

Data de hoje: {today.strftime('%d/%m/%Y')}

SUA TAREFA: Buscar passagens saindo de SALVADOR (SSA) para destinos no Brasil, Europa e EUA nas janelas:
- 30 dias: até {d30.strftime('%d/%m/%Y')}
- 60 dias: até {d60.strftime('%d/%m/%Y')}
- 90 dias: até {d90.strftime('%d/%m/%Y')}

Critérios:
- Viagens de 4 a 7 dias, próximas a finais de semana (quinta a domingo ou sexta a terça)
- Priorize voos mais baratos em reais, mas evite conexões inviáveis (acima de 4h de escala)
- Considere flexibilidade de ±2 dias nas datas
- Compare múltiplos destinos e companhias aéreas

Responda APENAS com JSON válido, sem texto fora do JSON, sem markdown, sem blocos de código. Estrutura exata:

{{
  "data_geracao": "{today.strftime('%d/%m/%Y')}",
  "voos": [
    {{
      "destino": "string",
      "janela_dias": 30,
      "datas_sugeridas": "string",
      "companhia": "string",
      "preco_total_reais": "string",
      "tempo_viagem": "string",
      "escalas": "string",
      "praticidade": 4,
      "link_compra": "string",
      "observacoes": "string"
    }}
  ],
  "ranking_top3_baratos": [
    {{"posicao": 1, "destino": "string", "preco": "string", "motivo": "string"}}
  ],
  "melhor_opcao_geral": {{
    "brasil": {{"destino": "string", "preco": "string", "motivo": "string", "link": "string"}},
    "europa": {{"destino": "string", "preco": "string", "motivo": "string", "link": "string"}},
    "eua":    {{"destino": "string", "preco": "string", "motivo": "string", "link": "string"}}
  }},
  "insights": {{
    "melhor_janela":  "string",
    "dicas":          ["string", "string", "string"],
    "comprar_agora":  "string"
  }}
}}

Inclua pelo menos 12 destinos variados (Brasil, Europa, EUA). Use preços reais e atuais baseados no mercado de hoje."""

# ── GEMINI API ────────────────────────────────────────────────────────────────
def call_gemini(prompt: str) -> dict:
    url = (
        f"https://generativelanguage.googleapis.com/v1beta/models/"
        f"{MODEL}:generateContent?key={GEMINI_API_KEY}"
    )
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.3, "maxOutputTokens": 8192}
    }
    r = requests.post(url, json=payload, timeout=120)
    r.raise_for_status()
    raw = r.json()["candidates"][0]["content"]["parts"][0]["text"]
    # Strip possible markdown fences
    raw = raw.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
    return json.loads(raw)

# ── EXCEL ─────────────────────────────────────────────────────────────────────
DARK_BROWN = "2C1A0E"
GOLD       = "C8922A"
LIGHT_GOLD = "F5D68A"
WHITE      = "FFFFFF"
DARK_GRAY  = "3D3D3D"
GREEN_DARK = "1A6B35"
PALE_GOLD  = "FFF8E7"

def hdr_cell(c, text, bg=GOLD, fg=WHITE, size=10, bold=True):
    c.value      = text
    c.font       = Font(bold=bold, size=size, color=fg, name="Arial")
    c.fill       = PatternFill("solid", fgColor=bg)
    c.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    c.border     = Border(left=thin, right=thin, top=thin, bottom=thin)

def data_cell(c, text, bg=WHITE, bold=False, align="left", color=DARK_GRAY, size=9):
    c.value     = text
    c.font      = Font(bold=bold, size=size, color=color, name="Arial")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    c.border    = Border(left=thin, right=thin, top=thin, bottom=thin)

def build_excel(data: dict, filepath: str):
    wb   = Workbook()
    today_str = data.get("data_geracao", datetime.date.today().strftime("%d/%m/%Y"))

    # ── Aba 1: Comparativo ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "✈ Comparativo"
    ws1.freeze_panes = "A4"

    ws1.merge_cells("A1:J1")
    c = ws1["A1"]
    c.value     = f"✈  PASSAGENS DE SALVADOR (SSA)  —  Gerado em {today_str}"
    c.font      = Font(bold=True, size=14, color=WHITE, name="Arial")
    c.fill      = PatternFill("solid", fgColor=DARK_BROWN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 34

    ws1.merge_cells("A2:J2")
    c = ws1["A2"]
    c.value     = "Janelas: 30 · 60 · 90 dias  |  Duração: 4–7 dias  |  Próximos a fins de semana"
    c.font      = Font(italic=True, size=10, color=DARK_BROWN, name="Arial")
    c.fill      = PatternFill("solid", fgColor=LIGHT_GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 18

    headers = ["Destino", "Janela\n(dias)", "Datas Sugeridas", "Cia. Aérea",
               "Preço Total\n(R$)", "Tempo\nViagem", "Escalas",
               "Praticidade\n⭐", "Link de Compra", "Observações"]
    ws1.row_dimensions[3].height = 38
    for col, h in enumerate(headers, 1):
        hdr_cell(ws1.cell(3, col), h)

    row_fills = [WHITE, PALE_GOLD]
    for i, v in enumerate(data.get("voos", []), 4):
        bg = row_fills[i % 2]
        ws1.row_dimensions[i].height = 52
        vals = [
            v.get("destino",""), v.get("janela_dias",""), v.get("datas_sugeridas",""),
            v.get("companhia",""), v.get("preco_total_reais",""), v.get("tempo_viagem",""),
            v.get("escalas",""), f"{'⭐'*int(v.get('praticidade',3))}",
            v.get("link_compra",""), v.get("observacoes","")
        ]
        for col, val in enumerate(vals, 1):
            cc = ws1.cell(i, col, val)
            if col == 5:
                data_cell(cc, val, bg=bg, bold=True, align="center", color=GREEN_DARK)
            elif col == 9:
                data_cell(cc, val, bg=bg, size=8, color="0563C1")
                cc.font = Font(size=8, name="Arial", color="0563C1", underline="single")
            elif col in (1,2,6,7,8):
                data_cell(cc, val, bg=bg, align="center")
            else:
                data_cell(cc, val, bg=bg)

    col_widths = [24,10,24,20,16,14,16,12,42,52]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Aba 2: Ranking ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("🏆 Ranking")
    ws2.merge_cells("A1:D1")
    c = ws2["A1"]
    c.value     = "🏆  RANKING — TOP 3 MAIS BARATOS"
    c.font      = Font(bold=True, size=13, color=WHITE, name="Arial")
    c.fill      = PatternFill("solid", fgColor=DARK_BROWN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    for col, h in enumerate(["#","Destino","Preço (R$)","Por quê?"], 1):
        hdr_cell(ws2.cell(3, col), h)
    ws2.row_dimensions[3].height = 22

    medals = ["FFF0D0","F0F0F0","FFF8EE"]
    for i, r in enumerate(data.get("ranking_top3_baratos",[]), 4):
        ws2.row_dimensions[i].height = 30
        for col, val in enumerate([r.get("posicao",""), r.get("destino",""),
                                    r.get("preco",""), r.get("motivo","")], 1):
            data_cell(ws2.cell(i,col), str(val), bg=medals[i-4],
                      align="center" if col!=4 else "left", bold=(col==1))

    ws2.merge_cells("A6:D6")
    c = ws2["A6"]
    c.value     = "⭐  MELHOR OPÇÃO GERAL POR CATEGORIA"
    c.font      = Font(bold=True, size=11, color=WHITE, name="Arial")
    c.fill      = PatternFill("solid", fgColor=GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[6].height = 24

    for col, h in enumerate(["Categoria","Destino","Preço (R$)","Motivo / Link"], 1):
        hdr_cell(ws2.cell(7, col), h)

    best = data.get("melhor_opcao_geral", {})
    cat_map = [("🇧🇷 Brasil","brasil"),("🌍 Europa","europa"),("🇺🇸 EUA","eua")]
    best_fills = [WHITE, PALE_GOLD, WHITE]
    for i, (label, key) in enumerate(cat_map, 8):
        d = best.get(key, {})
        ws2.row_dimensions[i].height = 36
        row_vals = [label, d.get("destino",""), d.get("preco",""),
                    f"{d.get('motivo','')}  |  {d.get('link','')}"]
        for col, val in enumerate(row_vals, 1):
            data_cell(ws2.cell(i,col), val, bg=best_fills[i-8],
                      align="center" if col in (1,2,3) else "left")

    for col, w in zip("ABCD",[16,24,16,60]):
        ws2.column_dimensions[col].width = w

    # ── Aba 3: Insights ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("💡 Insights")
    ws3.merge_cells("A1:B1")
    c = ws3["A1"]
    c.value     = "💡  INSIGHTS ESTRATÉGICOS DO DIA"
    c.font      = Font(bold=True, size=13, color=WHITE, name="Arial")
    c.fill      = PatternFill("solid", fgColor=DARK_BROWN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 32

    ins = data.get("insights", {})
    insight_rows = [
        ("🗓 Melhor janela para comprar agora", ins.get("melhor_janela","")),
        ("⚡ Vale esperar ou comprar hoje?",    ins.get("comprar_agora","")),
    ]
    for i, dica in enumerate(ins.get("dicas",[]), 1):
        insight_rows.append((f"💡 Dica {i}", dica))

    for i, (label, val) in enumerate(insight_rows, 3):
        ws3.row_dimensions[i].height = 44
        c1 = ws3.cell(i, 1, label)
        c1.font      = Font(bold=True, size=10, name="Arial", color=DARK_BROWN)
        c1.fill      = PatternFill("solid", fgColor=LIGHT_GOLD)
        c1.alignment = Alignment(vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        c1.border    = Border(left=thin,right=thin,top=thin,bottom=thin)

        c2 = ws3.cell(i, 2, val)
        data_cell(c2, val, bg=(WHITE if i%2==0 else PALE_GOLD))

    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 90

    wb.save(filepath)

# ── E-MAIL ────────────────────────────────────────────────────────────────────
def send_email(filepath: str, today_str: str):
    msg = MIMEMultipart()
    msg["From"]    = GMAIL_USER
    msg["To"]      = ", ".join(TO_EMAILS)
    msg["Subject"] = f"✈ Passagens de Salvador — Relatório {today_str}"

    body = f"""
Olá! 👋

Segue em anexo o relatório diário de passagens saindo de Salvador (SSA) para Brasil, Europa e EUA.

📅 Data: {today_str}
📊 Conteúdo: Comparativo completo · Ranking Top 3 · Insights estratégicos
🗓 Janelas analisadas: 30, 60 e 90 dias

Abra a planilha Excel para ver todos os detalhes, links de compra e recomendações do dia.

Boas viagens! ✈
    """.strip()

    msg.attach(MIMEText(body, "plain", "utf-8"))

    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        filename = f"passagens_ssa_{today_str.replace('/','')}.xlsx"
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_PASSWORD)
        server.sendmail(GMAIL_USER, TO_EMAILS, msg.as_string())

    print(f"✅ E-mail enviado para: {', '.join(TO_EMAILS)}")

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    today_str = datetime.date.today().strftime("%d/%m/%Y")
    print(f"🔍 Consultando Gemini ({MODEL})...")

    prompt = build_prompt()
    data   = call_gemini(prompt)
    print(f"✅ Recebidos {len(data.get('voos',[]))} voos do Gemini")

    filepath = f"/tmp/passagens_ssa_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    print("📊 Gerando planilha Excel...")
    build_excel(data, filepath)
    print(f"✅ Planilha salva: {filepath}")

    print("📧 Enviando e-mail...")
    send_email(filepath, today_str)

if __name__ == "__main__":
    main()
